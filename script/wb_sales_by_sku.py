#!/usr/bin/env python3
"""
读取 WB 供应商销量报告，按 SKU 汇总已下单数（L列）和已签收数（N列），
并与 ec_orders 数据库对比，输出差值。

用法：
    python3 script/wb_sales_by_sku.py <xlsx路径> [account_id] [from_date] [to_date]

示例：
    python3 script/wb_sales_by_sku.py ~/Downloads/supplier-goods-xxx.XLSX 3 2025-01-01 2026-05-31
    python3 script/wb_sales_by_sku.py ~/Downloads/supplier-goods-xxx.XLSX 2

account_id 默认 3（МИРОВОЙ ВЫБОР）；日期默认 2025-01-01 ~ 2026-05-31，莫斯科时区。
文件格式：第1行标题，第2行表头，第3行起数据；
  F列 = 卖家货号(offer_id)，L列 = 已下单件，N列 = 已签收件。
"""

import openpyxl
import json
import subprocess
import sys
from pathlib import Path
from collections import defaultdict

ACCOUNT_ID   = 3
FROM_DATE    = "2025-01-01"
TO_DATE      = "2026-05-31"
CONTAINER_CMD = "docker ps --format '{{.Names}}' | grep eshop_manage-web | head -1"

# ─── 1. 读取 xlsx ─────────────────────────────────────────────────────────────

def load_xlsx(path):
    """返回 { offer_id: { ordered: int, delivered: int } }"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    totals = defaultdict(lambda: {"ordered": 0, "delivered": 0})
    skipped = 0

    for row in ws.iter_rows(min_row=3, values_only=True):
        offer_id  = row[5]   # F列（0-indexed=5）
        ordered   = row[11]  # L列
        delivered = row[13]  # N列

        if not offer_id:
            skipped += 1
            continue

        offer_id = str(offer_id).strip()
        totals[offer_id]["ordered"]   += int(ordered   or 0)
        totals[offer_id]["delivered"] += int(delivered or 0)

    print(f"  读取完成：{len(totals)} 个 offer_id，跳过空行 {skipped} 条")
    return dict(totals)


# ─── 2. Rails runner 模板 ─────────────────────────────────────────────────────

RAILS_TEMPLATE = r"""
$stdout.sync = true
Rails.logger = Logger.new(IO::NULL)

ACCOUNT_ID = ACCOUNT_ID_PLACEHOLDER
FROM_DATE  = 'FROM_DATE_PLACEHOLDER'
TO_DATE    = 'TO_DATE_PLACEHOLDER'
data       = JSON.parse('DATA_JSON_PLACEHOLDER')  # { offer_id => { ordered, delivered } }

store = Ec::Store.find_by(platform: 'wb', wb_raw_account_id: ACCOUNT_ID)
unless store
  puts "❌ 找不到 Ec::Store (platform=wb, wb_raw_account_id=#{ACCOUNT_ID})"
  exit 1
end
puts "Store: #{store.store_name} (id=#{store.id})  #{FROM_DATE} ~ #{TO_DATE} (MSK)"

# ── Section 1: xlsx 汇总 ─────────────────────────────────────────────────────

mapping   = Ec::SkuProduct
  .where(store_id: store.id, offer_id: data.keys)
  .pluck(:offer_id, :sku_code)
  .to_h
unmatched = data.keys - mapping.keys

xlsx_by_sku = Hash.new { |h, k| h[k] = { ordered: 0, delivered: 0 } }
data.each do |offer_id, nums|
  sku = mapping[offer_id] || "(未匹配) #{offer_id}"
  xlsx_by_sku[sku][:ordered]   += nums["ordered"].to_i
  xlsx_by_sku[sku][:delivered] += nums["delivered"].to_i
end

# ── Section 2: ec_orders 查询（莫斯科时区） ──────────────────────────────────

# 从 ec_sku_products 拿全部匹配 id，构建 { 任意id => sku_code } 的反向映射
sku_products = Ec::SkuProduct.where(store_id: store.id, platform: 'wb')
id_to_sku = {}
sku_products.each do |sp|
  [sp.sku_code, sp.product_id, sp.offer_id].compact_blank.each do |id|
    id_to_sku[id.to_s] ||= sp.sku_code
  end
end

all_ids      = id_to_sku.keys
sku_code_set = sku_products.pluck(:sku_code).compact.uniq

# 条件：sku_code 直接命中，或 offer_id/platform_sku_id 命中
tz_expr = "(ec_orders.ordered_at AT TIME ZONE 'Europe/Moscow')::date"

base = Ec::OrderItem
  .joins(:order)
  .joins("INNER JOIN ec_order_fulfillments ef ON ef.order_id = ec_orders.id AND ef.fulfillment_type = 'fbs'")
  .where(ec_order_items: { platform: 'wb', store_id: store.id })
  .where("#{tz_expr} BETWEEN ? AND ?", FROM_DATE, TO_DATE)
  .where(
    "ec_order_items.sku_code IN (?) OR ec_order_items.offer_id IN (?) OR ec_order_items.platform_sku_id IN (?)",
    sku_code_set, all_ids, all_ids
  )

# 拉原始行，在 Ruby 侧按 sku_code 重新归组
rows_ordered   = base.pluck("ec_order_items.sku_code", "ec_order_items.offer_id",
                             "ec_order_items.platform_sku_id", "ec_order_items.quantity")
rows_delivered = base.where.not(ec_orders: { order_status: %w[cancelled returned] })
                     .pluck("ec_order_items.sku_code", "ec_order_items.offer_id",
                            "ec_order_items.platform_sku_id", "ec_order_items.quantity")

def resolve_sku(sku_code, offer_id, platform_sku_id, id_to_sku)
  id_to_sku[sku_code.to_s] || id_to_sku[offer_id.to_s] || id_to_sku[platform_sku_id.to_s]
end

ec_ordered   = Hash.new(0)
rows_ordered.each do |sc, oid, psid, qty|
  sku = resolve_sku(sc, oid, psid, id_to_sku)
  ec_ordered[sku] += qty.to_i if sku
end

ec_delivered = Hash.new(0)
rows_delivered.each do |sc, oid, psid, qty|
  sku = resolve_sku(sc, oid, psid, id_to_sku)
  ec_delivered[sku] += qty.to_i if sku
end

# ── Section 3: 打印对比 ──────────────────────────────────────────────────────

all_skus = (xlsx_by_sku.keys + ec_ordered.keys + ec_delivered.keys)
  .compact
  .reject { |s| s.start_with?("(未匹配)") }
  .uniq.sort

sep = "=" * 90
puts "\n" + sep
puts "  WB FBS 销量对比  store=#{store.store_name}  #{FROM_DATE} ~ #{TO_DATE} (MSK)"
puts sep
hdr = "  #{"SKU".ljust(20)}  #{"xlsx下单".rjust(8)}  #{"ec下单".rjust(8)}  #{"差值".rjust(6)}  |  #{"xlsx签收".rjust(8)}  #{"ec签收".rjust(8)}  #{"差值".rjust(6)}"
puts hdr
puts "  " + "-" * 86

tot = { xo: 0, eo: 0, xd: 0, ed: 0 }
all_skus.each do |sku|
  xo = xlsx_by_sku[sku]&.dig(:ordered)   || 0
  xd = xlsx_by_sku[sku]&.dig(:delivered) || 0
  eo = ec_ordered[sku]   || 0
  ed = ec_delivered[sku] || 0
  do_ = eo - xo
  dd  = ed - xd
  flag_o = do_.abs > 0 ? (do_ > 0 ? " ▲" : " ▼") : "  "
  flag_d = dd.abs  > 0 ? (dd  > 0 ? " ▲" : " ▼") : "  "
  puts "  #{sku.ljust(20)}  #{xo.to_s.rjust(8)}  #{eo.to_s.rjust(8)}  #{(do_ >= 0 ? "+#{do_}" : do_.to_s).rjust(6)}#{flag_o}  |  #{xd.to_s.rjust(8)}  #{ed.to_s.rjust(8)}  #{(dd >= 0 ? "+#{dd}" : dd.to_s).rjust(6)}#{flag_d}"
  tot[:xo] += xo; tot[:eo] += eo; tot[:xd] += xd; tot[:ed] += ed
end

puts "  " + "-" * 86
do_t = tot[:eo] - tot[:xo]; dd_t = tot[:ed] - tot[:xd]
puts "  #{"合计".ljust(20)}  #{tot[:xo].to_s.rjust(8)}  #{tot[:eo].to_s.rjust(8)}  #{(do_t >= 0 ? "+#{do_t}" : do_t.to_s).rjust(6)}     |  #{tot[:xd].to_s.rjust(8)}  #{tot[:ed].to_s.rjust(8)}  #{(dd_t >= 0 ? "+#{dd_t}" : dd_t.to_s).rjust(6)}"

unless unmatched.empty?
  puts "\n  ⚠️  未匹配 offer_id（#{unmatched.size} 个，未计入上表）："
  unmatched.each { |o| puts "      #{o}  →  xlsx下单=#{data[o]['ordered']}  xlsx签收=#{data[o]['delivered']}" }
end

# ec_orders 里有但 xlsx 无的 sku
ec_only = (ec_ordered.keys + ec_delivered.keys).compact.uniq.sort - all_skus - [""]
unless ec_only.empty?
  puts "\n  ℹ️  仅在 ec_orders 中出现的 SKU（#{ec_only.size} 个）："
  ec_only.each { |s| puts "      #{s}  →  ec下单=#{ec_ordered[s] || 0}  ec签收=#{ec_delivered[s] || 0}" }
end

# raw_wb_supply_items 总签收（nm_id → sku_code）
nm_id_to_sku = Ec::SkuProduct
  .where(store_id: store.id, platform: 'wb')
  .where.not(product_id: [nil, ''])
  .pluck(:product_id, :sku_code)
  .to_h

supply_by_sku = RawWb::SupplyItem
  .joins("INNER JOIN raw_wb_supplies ON raw_wb_supplies.wb_supply_id = raw_wb_supply_items.wb_supply_id")
  .where(raw_wb_supply_items: { account_id: ACCOUNT_ID, nm_id: nm_id_to_sku.keys.map(&:to_i) })
  .where("raw_wb_supplies.account_id = ?", ACCOUNT_ID)
  .where("raw_wb_supplies.supply_created_at <= ?", "#{TO_DATE} 23:59:59")
  .group("raw_wb_supply_items.nm_id")
  .sum("raw_wb_supply_items.accepted_qty")
  .transform_keys { |nm| nm_id_to_sku[nm.to_s] }

# JSON 数据块供 Python 生成 xlsx（差值为负的行）
rows_json = all_skus.map do |sku|
  xo = xlsx_by_sku[sku]&.dig(:ordered)   || 0
  xd = xlsx_by_sku[sku]&.dig(:delivered) || 0
  eo = ec_ordered[sku]   || 0
  ed = ec_delivered[sku] || 0
  next nil unless xo > eo  # 只保留差值为负（xlsx > ec）的行
  { sku: sku, xlsx_ordered: xo, ec_ordered: eo, fbw_ordered: xo - eo,
    xlsx_delivered: xd, ec_delivered: ed, fbw_delivered: xd - ed,
    supply_accepted: supply_by_sku[sku] || 0 }
end.compact
puts "\nJSON_DATA_START"
puts rows_json.to_json
puts "JSON_DATA_END"
"""

def build_rails_script(totals, account_id, from_date, to_date):
    script = RAILS_TEMPLATE
    script = script.replace("ACCOUNT_ID_PLACEHOLDER", str(account_id))
    script = script.replace("FROM_DATE_PLACEHOLDER",  from_date)
    script = script.replace("TO_DATE_PLACEHOLDER",    to_date)
    script = script.replace("DATA_JSON_PLACEHOLDER",  json.dumps(totals, ensure_ascii=False))
    return script


# ─── 3. 生成差值 xlsx ────────────────────────────────────────────────────────

def save_xlsx(rows, xlsx_path):
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FBW差值"

    headers = ["SKU", "总下单", "FBS总下单", "FBW总下单", "总签收", "FBS总签收", "FBW总签收", "送仓总签收"]
    keys    = ["sku", "xlsx_ordered", "ec_ordered", "fbw_ordered",
               "xlsx_delivered", "ec_delivered", "fbw_delivered", "supply_accepted"]

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(keys, 1):
            ws.cell(row=row_idx, column=col_idx, value=row[key])

    # 合计行
    total_row = len(rows) + 2
    ws.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
    for col_idx, key in enumerate(keys[1:], 2):
        ws.cell(row=total_row, column=col_idx, value=sum(r[key] for r in rows)).font = Font(bold=True)

    col_widths = [22, 10, 12, 12, 10, 12, 12, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    wb.save(xlsx_path)
    print(f"xlsx 已保存到：{xlsx_path}")


# ─── 4. 上传并在服务器执行 ────────────────────────────────────────────────────

def get_container():
    result = subprocess.run(
        ["ssh", "root@eshop.evexport.cn", CONTAINER_CMD],
        capture_output=True, text=True
    )
    return result.stdout.strip()

def run_on_server(rails_script, out_path):
    container = get_container()
    if not container:
        print("❌ 无法获取容器名", file=sys.stderr)
        sys.exit(1)
    print(f"容器：{container}\n")

    local_path  = "/tmp/wb_sales_by_sku.rb"
    remote_path = "/tmp/wb_sales_by_sku.rb"

    with open(local_path, "w") as f:
        f.write(rails_script)

    subprocess.run(["scp", local_path, f"root@eshop.evexport.cn:{remote_path}"], check=True)
    subprocess.run(["ssh", "root@eshop.evexport.cn",
                    f"docker cp {remote_path} {container}:{remote_path}"], check=True)

    result = subprocess.run(
        ["ssh", "root@eshop.evexport.cn",
         f"docker exec {container} bin/rails runner {remote_path}"],
        capture_output=True, text=True
    )

    output = "\n".join(
        line for line in result.stdout.splitlines()
        if "image_processing" not in line and "image variants" not in line
    )

    # 分离 JSON 数据块和人类可读输出
    lines = result.stdout.splitlines()
    json_lines, display_lines, in_json = [], [], False
    for line in lines:
        if "image_processing" in line or "image variants" in line:
            continue
        if line.strip() == "JSON_DATA_START":
            in_json = True
        elif line.strip() == "JSON_DATA_END":
            in_json = False
        elif in_json:
            json_lines.append(line)
        else:
            display_lines.append(line)

    output = "\n".join(display_lines)
    print(output)
    if result.returncode != 0:
        print(result.stderr, file=sys.stderr)

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(output)
    print(f"\n结果已保存到：{out_path}")

    if json_lines:
        rows = json.loads("".join(json_lines))
        save_xlsx(rows, out_path.with_suffix(".xlsx"))


# ─── main ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    path       = Path(sys.argv[1]).expanduser()
    account_id = int(sys.argv[2])  if len(sys.argv) >= 3 else ACCOUNT_ID
    from_date  = sys.argv[3]       if len(sys.argv) >= 4 else FROM_DATE
    to_date    = sys.argv[4]       if len(sys.argv) >= 5 else TO_DATE
    out_path   = path.parent / (path.stem + "_by_sku.txt")

    print(f"\n读取文件：{path}  account_id={account_id}  {from_date}~{to_date}")
    totals = load_xlsx(path)

    rails_script = build_rails_script(totals, account_id, from_date, to_date)
    run_on_server(rails_script, out_path)
